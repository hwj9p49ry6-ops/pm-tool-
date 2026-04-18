import json
import os
import sys
import sqlite3
import io
import hashlib
import secrets

from flask import Flask, jsonify, request, render_template, g, send_file

app = Flask(__name__, template_folder='templates')
app.config['DATABASE'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pm.db')
app.secret_key = 'dev-secret-key'


@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()


# ── Override get_db to use app-relative path ─────────────────────────────────
import sqlite3

def _get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(app.config['DATABASE'])
        db.row_factory = sqlite3.Row
        db.execute("PRAGMA foreign_keys = ON")
    return db


# ── Serve SPA ─────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')


# ── Projects ──────────────────────────────────────────────────────────────────
@app.route('/api/projects', methods=['GET'])
def list_projects():
    db = _get_db()
    rows = db.execute("SELECT * FROM projects ORDER BY created_at").fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/projects', methods=['POST'])
def create_project():
    data = request.get_json() or {}
    name = data.get('name', 'New Project')
    color = data.get('color', '#4A90E2')
    db = _get_db()
    cur = db.execute("INSERT INTO projects (name, color) VALUES (?, ?)", (name, color))
    db.commit()
    row = db.execute("SELECT * FROM projects WHERE id=?", (cur.lastrowid,)).fetchone()
    return jsonify(dict(row)), 201


@app.route('/api/projects/<int:pid>', methods=['PUT'])
def update_project(pid):
    data = request.get_json() or {}
    db = _get_db()
    row = db.execute("SELECT * FROM projects WHERE id=?", (pid,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    name = data.get('name', row['name'])
    color = data.get('color', row['color'])
    db.execute("UPDATE projects SET name=?, color=? WHERE id=?", (name, color, pid))
    db.commit()
    row = db.execute("SELECT * FROM projects WHERE id=?", (pid,)).fetchone()
    return jsonify(dict(row))


@app.route('/api/projects/<int:pid>', methods=['DELETE'])
def delete_project(pid):
    db = _get_db()
    row = db.execute("SELECT id FROM projects WHERE id=?", (pid,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    db.execute("DELETE FROM projects WHERE id=?", (pid,))
    db.commit()
    return jsonify({'ok': True})


# ── Tasks ─────────────────────────────────────────────────────────────────────
@app.route('/api/projects/<int:pid>/tasks', methods=['GET'])
def list_tasks(pid):
    db = _get_db()
    if not db.execute("SELECT id FROM projects WHERE id=?", (pid,)).fetchone():
        return jsonify({'error': 'Not found'}), 404
    rows = db.execute(
        "SELECT * FROM tasks WHERE project_id=? ORDER BY sort_order, id",
        (pid,)
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/projects/<int:pid>/tasks', methods=['POST'])
def create_task(pid):
    data = request.get_json() or {}
    db = _get_db()
    if not db.execute("SELECT id FROM projects WHERE id=?", (pid,)).fetchone():
        return jsonify({'error': 'Not found'}), 404
    # Determine next sort_order
    max_order = db.execute(
        "SELECT COALESCE(MAX(sort_order),0) FROM tasks WHERE project_id=?", (pid,)
    ).fetchone()[0]
    cur = db.execute(
        """INSERT INTO tasks
           (project_id, name, owner, status, priority, start_date, end_date,
            pct_complete, depends_on, notes, sort_order, parent_id)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
        (
            pid,
            data.get('name', 'New Task'),
            data.get('owner', ''),
            data.get('status', 'Not Started'),
            data.get('priority', 'Medium'),
            data.get('start_date', ''),
            data.get('end_date', ''),
            data.get('pct_complete', 0),
            data.get('depends_on', ''),
            data.get('notes', ''),
            data.get('sort_order', max_order + 10),
            data.get('parent_id', None),
        )
    )
    db.commit()
    row = db.execute("SELECT * FROM tasks WHERE id=?", (cur.lastrowid,)).fetchone()
    return jsonify(dict(row)), 201


@app.route('/api/tasks/<int:tid>', methods=['PUT'])
def update_task(tid):
    data = request.get_json() or {}
    db = _get_db()
    row = db.execute("SELECT * FROM tasks WHERE id=?", (tid,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404

    fields = ['name', 'owner', 'status', 'priority', 'start_date', 'end_date',
              'pct_complete', 'depends_on', 'notes', 'sort_order', 'duration',
              'radar_id', 'is_milestone', 'parent_id', 'color']
    updates = {}
    for f in fields:
        if f in data:
            updates[f] = data[f]

    if updates:
        set_clause = ', '.join(f"{k}=?" for k in updates)
        values = list(updates.values()) + [tid]
        db.execute(f"UPDATE tasks SET {set_clause} WHERE id=?", values)
        db.commit()

    row = db.execute("SELECT * FROM tasks WHERE id=?", (tid,)).fetchone()
    return jsonify(dict(row))


@app.route('/api/tasks/<int:tid>', methods=['DELETE'])
def delete_task(tid):
    db = _get_db()
    row = db.execute("SELECT id FROM tasks WHERE id=?", (tid,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    db.execute("UPDATE tasks SET parent_id=NULL WHERE parent_id=?", (tid,))
    db.execute("DELETE FROM tasks WHERE id=?", (tid,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/tasks/reorder', methods=['POST'])
def reorder_tasks():
    items = request.get_json() or []
    db = _get_db()
    for item in items:
        db.execute("UPDATE tasks SET sort_order=? WHERE id=?",
                   (item['sort_order'], item['id']))
    db.commit()
    return jsonify({'ok': True})


# ── Excel Export / Template ────────────────────────────────────────────────────
@app.route('/api/projects/<int:pid>/template/excel', methods=['GET'])
def template_excel(pid):
    """Export current tasks to Excel (editable). If no tasks, returns a blank template."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    import re as _re

    db = _get_db()
    proj = db.execute("SELECT * FROM projects WHERE id=?", (pid,)).fetchone()
    if not proj:
        return jsonify({'error': 'Project not found'}), 404

    raw_tasks = db.execute(
        "SELECT * FROM tasks WHERE project_id=? ORDER BY sort_order", (pid,)
    ).fetchall()
    tasks = [dict(t) for t in raw_tasks]

    # ── Build hierarchy display order (parents then children) ─────────────────
    child_map = {}
    roots = []
    for t in tasks:
        if t.get('parent_id'):
            child_map.setdefault(t['parent_id'], []).append(t)
        else:
            roots.append(t)

    display_list = []  # list of (task, level, is_parent)
    def _visit(t, level):
        children = child_map.get(t['id'], [])
        display_list.append((t, level, len(children) > 0))
        for c in children:
            _visit(c, level + 1)
    for t in roots:
        _visit(t, 0)

    # row number = 1-based position in display list
    task_to_row = {t['id']: i + 1 for i, (t, _, _) in enumerate(display_list)}

    def pred_to_display(depends_on):
        """Convert stored task-ID depends_on → row-number display format."""
        if not depends_on:
            return ''
        def _conv(m):
            tid = int(m.group(1))
            typ = (m.group(2) or 'FS').upper()
            lag = m.group(3) or ''
            return f"{task_to_row.get(tid, tid)}{typ}{lag}"
        return _re.sub(r'(\d+)(FS|SS)?([+-]\d+[dw])?', _conv, str(depends_on), flags=_re.IGNORECASE)

    # ── Column layout ──────────────────────────────────────────────────────────
    COLS = [
        ('Row #',          'row_num',      7,  'Reference — used for Parent Row # and Predecessor columns'),
        ('Task Name',      'name',         42, 'Task or phase name'),
        ('Owner',          'owner',        18, 'Person responsible'),
        ('Status',         'status',       16, 'Not Started / In Progress / Complete / Blocked'),
        ('Priority',       'priority',     12, 'Low / Medium / High / Critical'),
        ('Start Date',     'start_date',   14, 'YYYY-MM-DD'),
        ('Duration (days)','duration',     15, 'Working days'),
        ('End Date',       'end_date',     14, 'YYYY-MM-DD'),
        ('% Done',         'pct_complete',  9, '0 – 100'),
        ('Predecessor',    'depends_on',   16, 'Row # + type: 2FS, 3SS, 4FS+2d'),
        ('Parent Row #',   'parent_row',   13, 'Row # of parent task — blank = top-level'),
        ('Notes',          'notes',        36, 'Free text'),
        ('Radar ID',       'radar_id',     14, 'Apple Radar bug ID'),
        ('Task ID',        'task_id',      10, 'System ID — do not edit; used for updates on re-upload'),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = proj['name'][:31]

    # ── Styles ─────────────────────────────────────────────────────────────────
    hdr_fill  = PatternFill('solid', fgColor='1E2235')
    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    ref_fill  = PatternFill('solid', fgColor='EFF6FF')
    ref_font  = Font(color='1D4ED8', size=10, bold=True)
    par_fill  = PatternFill('solid', fgColor='E0F2FE')   # parent row highlight
    chd_fill  = PatternFill('solid', fgColor='F8FAFF')   # child row
    sys_fill  = PatternFill('solid', fgColor='F3F4F6')   # Task ID column
    sys_font  = Font(color='9CA3AF', size=9, italic=True)
    alt_fill  = PatternFill('solid', fgColor='F9FAFB')
    thin      = Side(style='thin', color='E0E4EC')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    STATUS_LIST   = '"Not Started,In Progress,Complete,Blocked"'
    PRIORITY_LIST = '"Low,Medium,High,Critical"'
    status_dv   = DataValidation(type='list', formula1=STATUS_LIST,   allow_blank=True, showDropDown=False)
    priority_dv = DataValidation(type='list', formula1=PRIORITY_LIST, allow_blank=True, showDropDown=False)
    ws.add_data_validation(status_dv)
    ws.add_data_validation(priority_dv)

    # ── Header row ─────────────────────────────────────────────────────────────
    for ci, (hdr, _, width, _tip) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'B2'

    # ── Data rows ──────────────────────────────────────────────────────────────
    if display_list:
        for i, (t, level, is_parent) in enumerate(display_list):
            ri  = i + 2
            rn  = i + 1  # 1-based Row #

            # Determine parent row number for this task
            parent_rn = task_to_row.get(t.get('parent_id')) if t.get('parent_id') else ''

            row_fill = par_fill if is_parent else (chd_fill if level > 0 else (alt_fill if ri % 2 == 0 else None))

            field_values = {
                'row_num':      rn,
                'name':         ('  ' * level + (t['name'] or '')),
                'owner':        t.get('owner') or '',
                'status':       t.get('status') or 'Not Started',
                'priority':     t.get('priority') or 'Medium',
                'start_date':   t.get('start_date') or '',
                'duration':     t.get('duration') or '',
                'end_date':     t.get('end_date') or '',
                'pct_complete': t.get('pct_complete') or 0,
                'depends_on':   pred_to_display(t.get('depends_on')),
                'parent_row':   parent_rn,
                'notes':        t.get('notes') or '',
                'radar_id':     t.get('radar_id') or '',
                'task_id':      t['id'],
            }

            for ci, (_, field, _, _) in enumerate(COLS, 1):
                cell = ws.cell(row=ri, column=ci, value=field_values.get(field, ''))
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=(field == 'notes'))

                if field == 'row_num':
                    cell.fill = ref_fill; cell.font = ref_font
                    cell.alignment = Alignment(horizontal='center', vertical='top')
                elif field == 'task_id':
                    cell.fill = sys_fill; cell.font = sys_font
                    cell.alignment = Alignment(horizontal='center', vertical='top')
                elif field == 'parent_row':
                    cell.fill = PatternFill('solid', fgColor='FEF9C3')
                    cell.alignment = Alignment(horizontal='center', vertical='top')
                elif field == 'name' and is_parent:
                    cell.font = Font(bold=True, size=11)
                    if row_fill: cell.fill = row_fill
                elif row_fill:
                    cell.fill = row_fill

                if field == 'status':
                    status_dv.add(cell)
                elif field == 'priority':
                    priority_dv.add(cell)

            ws.row_dimensions[ri].height = 18

    else:
        # No tasks yet — write 30 blank rows as a starter template
        for ri in range(2, 32):
            rn = ri - 1
            row_fill = alt_fill if ri % 2 == 0 else None
            for ci, (_, field, _, _) in enumerate(COLS, 1):
                cell = ws.cell(row=ri, column=ci)
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=(field == 'notes'))
                if row_fill: cell.fill = row_fill
                if field == 'row_num':
                    cell.value = rn
                    cell.fill = ref_fill; cell.font = ref_font
                    cell.alignment = Alignment(horizontal='center', vertical='top')
                elif field == 'parent_row':
                    cell.fill = PatternFill('solid', fgColor='FEF9C3')
                    cell.alignment = Alignment(horizontal='center', vertical='top')
                elif field == 'task_id':
                    cell.fill = sys_fill; cell.font = sys_font
                    cell.alignment = Alignment(horizontal='center', vertical='top')
                elif field == 'status':
                    cell.value = 'Not Started'; status_dv.add(cell)
                elif field == 'priority':
                    cell.value = 'Medium'; priority_dv.add(cell)
                elif field == 'pct_complete':
                    cell.value = 0

    # ── Instructions sheet ─────────────────────────────────────────────────────
    inst = wb.create_sheet('How To Use')
    inst.column_dimensions['A'].width = 90
    lines = [
        ('HOW TO USE THIS EXPORT', True, '1E2235'),
        ('', False, None),
        ('EDITING TASKS', True, '1D4ED8'),
        ('• Edit any cell and re-upload with the Upload button.', False, None),
        ('• Rows with a Task ID will be UPDATED. New rows (no Task ID) will be CREATED.', False, None),
        ('• Do NOT edit the Task ID or Row # columns.', False, None),
        ('', False, None),
        ('PARENT / CHILD HIERARCHY', True, '1D4ED8'),
        ('• Bold rows with blue background = parent tasks.', False, None),
        ('• To make a task a child, enter the parent task Row # in the "Parent Row #" column.', False, None),
        ('• Leave Parent Row # blank for top-level tasks.', False, None),
        ('', False, None),
        ('PREDECESSORS', True, '1D4ED8'),
        ('• Format: {Row #}{Type}  — FS = Finish-to-Start, SS = Start-to-Start.', False, None),
        ('• Optional lag: +Nd or +Nw  e.g. 3FS+2d', False, None),
        ('• Multiple predecessors: separate with comma, e.g. 2FS,4FS', False, None),
        ('', False, None),
        ('ADDING NEW TASKS', True, '1D4ED8'),
        ('• Add new rows at the bottom. Leave Task ID blank for new tasks.', False, None),
        ('• Set Row # sequentially. Use that Row # for Parent Row # and Predecessor references.', False, None),
    ]
    for row_i, (text, bold, color) in enumerate(lines, 1):
        cell = inst.cell(row=row_i, column=1, value=text)
        fnt = Font(bold=bold, size=11 if bold else 10)
        if color: fnt = Font(bold=bold, size=11 if bold else 10, color=color)
        cell.font = fnt

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    safe = ''.join(c if c.isalnum() or c in '-_ ' else '_' for c in proj['name'])
    suffix = '_export' if display_list else '_template'
    return send_file(buf, as_attachment=True,
                     download_name=f"{safe}{suffix}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



# ── Excel Import (template format — Row # based parent/predecessor) ────────────
@app.route('/api/projects/<int:pid>/import/excel', methods=['POST'])
def import_excel(pid):
    """Parse an uploaded Excel file (template or export) and create/update tasks."""
    from openpyxl import load_workbook
    import re as _re

    db = _get_db()
    if not db.execute("SELECT id FROM projects WHERE id=?", (pid,)).fetchone():
        return jsonify({'error': 'Project not found'}), 404

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    try:
        wb = load_workbook(io.BytesIO(f.read()), data_only=True)
    except Exception as e:
        return jsonify({'error': f'Could not read file: {e}'}), 400

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({'error': 'File is empty'}), 400

    headers = [str(h).strip() if h else '' for h in rows[0]]
    def col(name):
        try: return headers.index(name)
        except ValueError: return None

    def v(row, name, default=''):
        idx = col(name)
        if idx is None or idx >= len(row): return default
        val = row[idx]
        return val if val is not None else default

    def to_str(val):
        return str(val).strip() if val not in (None, '') else ''

    def to_date_str(val):
        """Normalize any Excel date value to YYYY-MM-DD string."""
        import datetime as _datetime
        if val is None or val == '':
            return ''
        if isinstance(val, (_datetime.datetime, _datetime.date)):
            return val.strftime('%Y-%m-%d')
        s = str(val).strip()
        if not s:
            return ''
        # Already YYYY-MM-DD (possibly with time suffix)
        if len(s) >= 10 and s[4] == '-' and s[7] == '-':
            return s[:10]
        # Try MM/DD/YYYY or M/D/YYYY
        for fmt in ('%m/%d/%Y', '%m/%d/%y', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                return _datetime.datetime.strptime(s, fmt).strftime('%Y-%m-%d')
            except ValueError:
                pass
        return ''  # unparseable — store blank rather than garbage

    VALID_STATUS   = {'Not Started','In Progress','Complete','Blocked'}
    VALID_PRIORITY = {'Low','Medium','High','Critical'}

    # Determine if this is a template file (has 'Row #') or legacy export (has 'ID')
    is_template = col('Row #') is not None

    # ── Pass 1: collect all data rows ─────────────────────────────────────────
    data_rows = []  # list of dicts
    for ri, row in enumerate(rows[1:], 1):
        name = to_str(v(row, 'Task Name *', '') or v(row, 'Task Name', ''))
        if not name or name.startswith('←'):  # skip blank + example rows
            continue

        status   = to_str(v(row, 'Status',   'Not Started')) or 'Not Started'
        priority = to_str(v(row, 'Priority', 'Medium'))      or 'Medium'
        if status   not in VALID_STATUS:   status   = 'Not Started'
        if priority not in VALID_PRIORITY: priority = 'Medium'

        try: duration = int(float(v(row,'Duration (days)',None) or v(row,'Duration (d)',None) or 0)) or None
        except: duration = None
        try: pct = max(0, min(100, int(float(v(row,'% Done',0) or 0))))
        except: pct = 0

        # Row # from the cell (template) or sequential index (legacy)
        if is_template:
            try: row_ref = int(v(row, 'Row #', ri))
            except: row_ref = ri
        else:
            row_ref = ri

        # Legacy export: task DB ID for updates
        try: task_id = int(v(row,'Task ID',None) or v(row,'ID',None) or 0) or None
        except: task_id = None

        data_rows.append({
            'row_ref':    row_ref,
            'task_id':    task_id,
            'name':       name,
            'owner':      to_str(v(row,'Owner','')),
            'status':     status,
            'priority':   priority,
            'start_date': to_date_str(v(row,'Start Date','')),
            'end_date':   to_date_str(v(row,'End Date','')),
            'duration':   duration,
            'pct':        pct,
            'depends_raw':to_str(v(row,'Predecessor','')),
            'parent_raw': to_str(v(row,'Parent Row #','') if is_template else v(row,'Parent ID','')),
            'notes':      to_str(v(row,'Notes','')),
            'radar_id':   to_str(v(row,'Radar ID','')),
        })

    if not data_rows:
        return jsonify({'error': 'No task rows found (all rows blank or example rows)'}), 400

    # ── Pass 2: create/update tasks, build row_ref → db_id map ───────────────
    row_ref_to_db_id = {}  # row_ref → real DB task id
    created = updated = skipped = 0
    errors  = []

    max_order = db.execute(
        "SELECT COALESCE(MAX(sort_order),0) FROM tasks WHERE project_id=?", (pid,)
    ).fetchone()[0]

    for i, dr in enumerate(data_rows):
        max_order += 10
        if dr['task_id']:
            # Legacy export update path
            existing = db.execute(
                "SELECT id FROM tasks WHERE id=? AND project_id=?", (dr['task_id'], pid)
            ).fetchone()
            if existing:
                db.execute("""UPDATE tasks SET name=?,owner=?,status=?,priority=?,
                    start_date=?,end_date=?,duration=?,pct_complete=?,
                    notes=?,radar_id=? WHERE id=?""",
                    (dr['name'],dr['owner'],dr['status'],dr['priority'],
                     dr['start_date'],dr['end_date'],dr['duration'],dr['pct'],
                     dr['notes'],dr['radar_id'],dr['task_id']))
                db.commit()
                row_ref_to_db_id[dr['row_ref']] = dr['task_id']
                updated += 1
            else:
                errors.append(f"Row {dr['row_ref']}: ID {dr['task_id']} not in project — creating as new")
                dr['task_id'] = None  # fall through to create

        if not dr['task_id']:
            cur = db.execute("""INSERT INTO tasks
                (project_id,name,owner,status,priority,start_date,end_date,
                 duration,pct_complete,notes,radar_id,sort_order)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                (pid,dr['name'],dr['owner'],dr['status'],dr['priority'],
                 dr['start_date'],dr['end_date'],dr['duration'],dr['pct'],
                 dr['notes'],dr['radar_id'],max_order))
            db.commit()
            row_ref_to_db_id[dr['row_ref']] = cur.lastrowid
            created += 1

    # ── Pass 3: wire parent_id + depends_on using row_ref → db_id map ─────────
    for dr in data_rows:
        db_id = row_ref_to_db_id.get(dr['row_ref'])
        if not db_id:
            continue
        updates = {}

        # Parent: translate row_ref → db_id
        if dr['parent_raw']:
            try:
                p_ref = int(float(dr['parent_raw']))
                p_db  = row_ref_to_db_id.get(p_ref)
                if p_db and p_db != db_id:
                    updates['parent_id'] = p_db
                else:
                    errors.append(f"Row {dr['row_ref']}: Parent Row # {p_ref} not found — task created as top-level")
            except (ValueError, TypeError):
                pass

        # Predecessor: translate each row_ref to db_id
        if dr['depends_raw']:
            def translate_pred(m):
                ref = int(m.group(1))
                typ = m.group(2) or 'FS'
                lag = m.group(3) or ''
                db_pred = row_ref_to_db_id.get(ref, ref)
                return f"{db_pred}{typ}{lag}"
            updates['depends_on'] = _re.sub(
                r'(\d+)(FS|SS)?([+-]\d+[dw])?', translate_pred,
                dr['depends_raw'], flags=_re.IGNORECASE
            )

        if updates:
            set_clause = ', '.join(f"{k}=?" for k in updates)
            db.execute(f"UPDATE tasks SET {set_clause} WHERE id=?",
                       list(updates.values()) + [db_id])

    db.commit()
    return jsonify({'ok':True,'created':created,'updated':updated,'skipped':skipped,'errors':errors})



@app.route('/api/radar/<radar_id>', methods=['GET'])
def get_radar_info(radar_id):
    """Fetch key Radar fields via MCP radar tools (subprocess)."""
    try:
        import subprocess
        result = subprocess.run(
            ['python3', os.path.join(_ROOT, 'integrations/radar/radar.py'),
             'get_radar', radar_id],
            capture_output=True, text=True, timeout=15,
            cwd=_ROOT
        )
        output = result.stdout.strip()
        if not output:
            return jsonify({'error': result.stderr or 'No output from radar'}), 502
        # radar.py may print a validation line before JSON
        start = output.find('{')
        if start >= 0:
            return jsonify(json.loads(output[start:]))
        return jsonify({'error': 'Unexpected radar output', 'raw': output[:300]}), 502
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def _run_radar(args, timeout=20):
    """Run radar.py CLI and return parsed JSON result."""
    import subprocess
    cmd = ['python3', os.path.join(_ROOT, 'integrations/radar/radar.py')] + args
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout, cwd=_ROOT)
    output = result.stdout.strip()
    if not output:
        raise RuntimeError(result.stderr or 'No output from radar')
    for start_char in ('{', '['):
        idx = output.find(start_char)
        if idx >= 0:
            return json.loads(output[idx:])
    raise RuntimeError(f'Unexpected radar output: {output[:200]}')


def _task_to_radar_description(task, parent_radar_id=None):
    """Build a Radar description from a task record."""
    lines = []
    if task.get('notes'):
        lines.append(task['notes'])
        lines.append('')
    lines.append('--- Schedule (PM Tool) ---')
    if task.get('start_date'):
        lines.append(f"Start Date: {task['start_date']}")
    if task.get('end_date'):
        lines.append(f"End Date:   {task['end_date']}")
    if task.get('duration'):
        lines.append(f"Duration:   {task['duration']}d")
    if task.get('owner'):
        lines.append(f"Owner:      {task['owner']}")
    lines.append(f"Status:     {task.get('status','Not Started')}")
    if task.get('pct_complete'):
        lines.append(f"Progress:   {task['pct_complete']}%")
    if parent_radar_id:
        lines.append(f"Parent Radar: rdar://{parent_radar_id}")
    return '\n'.join(lines)


def _task_priority_to_radar(priority):
    """Map task priority string to Radar priority integer."""
    return {'Critical': 1, 'High': 2, 'Medium': 3, 'Low': 4}.get(priority, 3)


def _task_status_to_radar_state(status):
    """Map task status to Radar state."""
    return {'Not Started': 'Analyze', 'In Progress': 'Build',
            'Blocked': 'Analyze', 'Complete': 'Closed'}.get(status, 'Analyze')


def _radar_state_to_task_status(state):
    """Map Radar state back to task status."""
    return {'Analyze': 'Not Started', 'Build': 'In Progress',
            'Integrate': 'In Progress', 'Verify': 'In Progress', 'Closed': 'Complete'}.get(state, 'In Progress')


@app.route('/api/projects/<int:pid>/generate-radars', methods=['POST'])
def generate_radars(pid):
    """
    Create Radars for all tasks that don't already have one.
    Body: { component: str, classification: str (default Task), skip_existing: bool }
    Returns: { created: N, skipped: N, errors: [...], results: [{task_id, task_name, radar_id}] }
    """
    data = request.get_json() or {}
    component = (data.get('component') or 'TV App').strip()
    classification = (data.get('classification') or 'Task').strip()
    skip_existing = data.get('skip_existing', True)

    db = _get_db()
    proj = db.execute("SELECT * FROM projects WHERE id=?", (pid,)).fetchone()
    if not proj:
        return jsonify({'error': 'Project not found'}), 404

    raw_tasks = db.execute(
        "SELECT * FROM tasks WHERE project_id=? ORDER BY sort_order, id", (pid,)
    ).fetchall()
    tasks = [dict(t) for t in raw_tasks]
    task_map = {t['id']: t for t in tasks}

    # Build hierarchy order: parents first, then children (depth-first)
    child_map = {}
    roots = []
    for t in tasks:
        pid_parent = t.get('parent_id')
        if pid_parent:
            child_map.setdefault(pid_parent, []).append(t)
        else:
            roots.append(t)

    ordered = []
    def _visit_task(t):
        ordered.append(t)
        for c in child_map.get(t['id'], []):
            _visit_task(c)
    for t in roots:
        _visit_task(t)

    # Map task_id → newly created radar_id (for wiring children to parent radars)
    task_to_new_radar = {}
    created = skipped = 0
    errors = []
    results = []

    for t in ordered:
        if skip_existing and t.get('radar_id'):
            task_to_new_radar[t['id']] = t['radar_id']
            skipped += 1
            continue

        # Determine parent radar ID (for description linkage)
        parent_task_id = t.get('parent_id')
        parent_radar_id = None
        if parent_task_id:
            parent_radar_id = task_to_new_radar.get(parent_task_id) or (task_map.get(parent_task_id) or {}).get('radar_id')

        description = _task_to_radar_description(t, parent_radar_id)
        priority = _task_priority_to_radar(t.get('priority', 'Medium'))
        title = (t.get('name') or 'Untitled Task').strip()

        # Build component string: if parent has a radar, append "sub-task of rdar://ID" in title
        full_title = title
        if parent_radar_id:
            full_title = f"{title} [sub: rdar://{parent_radar_id}]"

        try:
            cmd_args = [
                'create_radar',
                '--title', full_title,
                '--component', component,
                '--description', description,
                '--classification', classification,
                '--priority', str(priority),
            ]
            if t.get('owner'):
                cmd_args += ['--assignee', t['owner']]

            result = _run_radar(cmd_args, timeout=30)
            radar_id = None
            if isinstance(result, dict):
                radar_id = (result.get('radar') or {}).get('id') or result.get('id')

            if radar_id:
                db.execute("UPDATE tasks SET radar_id=? WHERE id=?", (str(radar_id), t['id']))
                db.commit()
                t['radar_id'] = str(radar_id)
                task_to_new_radar[t['id']] = str(radar_id)
                results.append({'task_id': t['id'], 'task_name': title, 'radar_id': str(radar_id)})
                created += 1
            else:
                errors.append({'task': title, 'error': f'No radar ID in response: {str(result)[:200]}'})
        except Exception as e:
            errors.append({'task': title, 'error': str(e)})

    return jsonify({'ok': True, 'created': created, 'skipped': skipped, 'errors': errors, 'results': results})


@app.route('/api/projects/<int:pid>/sync-radars', methods=['POST'])
def sync_radars(pid):
    """
    Bidirectional sync: for tasks with radar_id, pull state/assignee from Radar
    and push schedule (title, description with dates) back to Radar.
    Body: { direction: 'both' | 'pull' | 'push' }
    Returns: { synced: N, errors: [...], changes: [...] }
    """
    data = request.get_json() or {}
    direction = data.get('direction', 'both')

    db = _get_db()
    tasks = [dict(r) for r in db.execute(
        "SELECT * FROM tasks WHERE project_id=? AND radar_id != '' AND radar_id IS NOT NULL ORDER BY sort_order",
        (pid,)
    ).fetchall()]

    synced = 0
    errors = []
    changes = []

    for t in tasks:
        rid = t.get('radar_id', '').strip()
        if not rid:
            continue
        task_updates = {}

        try:
            # ── PULL: Radar → Task ─────────────────────────────────
            if direction in ('both', 'pull'):
                radar_data = _run_radar(['get_radar', rid], timeout=15)
                r = radar_data.get('radar', radar_data) if isinstance(radar_data, dict) else {}

                # State → Status
                radar_state = r.get('state') or r.get('status') or ''
                new_status = _radar_state_to_task_status(radar_state)
                if new_status and new_status != t.get('status') and radar_state:
                    task_updates['status'] = new_status
                    changes.append(f"[{t['name']}] status: {t['status']} → {new_status} (from Radar {radar_state})")

                # Assignee → Owner (only if task owner is blank)
                assignee = r.get('assignee') or r.get('assignee_email') or ''
                if assignee and not t.get('owner'):
                    task_updates['owner'] = assignee
                    changes.append(f"[{t['name']}] owner set to {assignee}")

                # Title → Name (only if Radar title was manually updated and differs)
                radar_title = r.get('title') or r.get('name') or ''
                # Don't clobber task name — only sync if task name was auto-set from radar
                # (identified by the task name matching a previously generated title pattern)

            # ── PUSH: Task → Radar ────────────────────────────────
            if direction in ('both', 'push'):
                parent_task = db.execute(
                    "SELECT radar_id FROM tasks WHERE id=?", (t.get('parent_id') or 0,)
                ).fetchone()
                parent_radar_id = dict(parent_task).get('radar_id') if parent_task else None
                description = _task_to_radar_description(t, parent_radar_id)

                push_payload = {
                    'title': t.get('name') or 'Untitled',
                    'description': description,
                }
                # Map task priority → radar priority
                prio_num = _task_priority_to_radar(t.get('priority', 'Medium'))
                push_payload['priority'] = str(prio_num)

                _run_radar([
                    'update_radar', rid,
                    '--title', push_payload['title'],
                    '--description', push_payload['description'],
                    '--priority', push_payload['priority'],
                ], timeout=15)
                changes.append(f"[{t['name']}] pushed to Radar {rid}")

            # Apply pulled updates to task
            if task_updates:
                set_clause = ', '.join(f"{k}=?" for k in task_updates)
                db.execute(f"UPDATE tasks SET {set_clause} WHERE id=?", list(task_updates.values()) + [t['id']])
                db.commit()

            synced += 1

        except Exception as e:
            errors.append({'task': t.get('name', rid), 'error': str(e)})

    return jsonify({'ok': True, 'synced': synced, 'errors': errors, 'changes': changes})


@app.route('/api/tasks/<int:tid>/push-to-radar', methods=['POST'])
def push_task_to_radar(tid):
    """Push a single task's current name/notes/dates to its linked Radar."""
    db = _get_db()
    row = db.execute("SELECT * FROM tasks WHERE id=?", (tid,)).fetchone()
    if not row:
        return jsonify({'error': 'Task not found'}), 404
    t = dict(row)
    rid = (t.get('radar_id') or '').strip()
    if not rid:
        return jsonify({'error': 'Task has no linked Radar'}), 400

    parent_row = db.execute(
        "SELECT radar_id FROM tasks WHERE id=?", (t.get('parent_id') or 0,)
    ).fetchone()
    parent_radar_id = dict(parent_row).get('radar_id') if parent_row else None

    description = _task_to_radar_description(t, parent_radar_id)
    prio = _task_priority_to_radar(t.get('priority', 'Medium'))

    try:
        result = _run_radar([
            'update_radar', rid,
            '--title', t.get('name') or 'Untitled',
            '--description', description,
            '--priority', str(prio),
        ], timeout=15)
        return jsonify({'ok': True, 'radar_id': rid, 'result': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500



@app.route('/api/radar/<radar_id>/sub-radars', methods=['GET'])
def get_sub_radars(radar_id):
    """Find sub-radars of a parent radar via radar.py CLI."""
    try:
        import subprocess
        result = subprocess.run(
            ['python3', os.path.join(_ROOT, 'integrations/radar/radar.py'),
             'find_radars', f'parent:{radar_id}'],
            capture_output=True, text=True, timeout=20,
            cwd=_ROOT
        )
        output = result.stdout.strip()
        if not output:
            return jsonify([])
        start = output.find('[')
        if start >= 0:
            return jsonify(json.loads(output[start:]))
        start = output.find('{')
        if start >= 0:
            data = json.loads(output[start:])
            return jsonify(data if isinstance(data, list) else [data])
        return jsonify([])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Keynote Export ────────────────────────────────────────────────────────────
@app.route('/api/projects/<int:pid>/export/keynote', methods=['POST'])
def export_keynote(pid):
    """Generate an Apple Keynote Black theme .pptx for this project."""
    import sys as _sys, io as _io
    _sys.path.insert(0, os.path.join(_ROOT, 'integrations', 'keynote'))
    try:
        from apple_keynote import (
            AppleKeynote, add_table_card, add_stat, add_risk_card,
            add_milestone_row, add_text_box, add_divider, add_tag,
            add_section_label, add_footer,
            MARGIN, MARGIN_TOP, SLIDE_W, SLIDE_H,
            TEXT_PRIMARY, TEXT_SECONDARY, TEXT_TERTIARY,
            BLUE, GREEN, RED, ORANGE, TEAL, INDIGO,
            FONT_BODY, FONT_MONO, FONT_DISPLAY,
            NEAR_BLACK, CARD_BG, CARD_BORDER, DIVIDER,
        )
        from pptx.util import Inches, Pt
        from pptx.dml.color import RGBColor
        from pptx.enum.text import PP_ALIGN
    except ImportError as e:
        return jsonify({'error': f'python-pptx not available: {e}'}), 500

    db = _get_db()
    proj = db.execute("SELECT * FROM projects WHERE id=?", (pid,)).fetchone()
    if not proj:
        return jsonify({'error': 'Project not found'}), 404

    data = request.get_json(silent=True) or {}
    opts = {
        'title':         data.get('title', proj['name']),
        'subtitle':      data.get('subtitle', ''),
        'author':        data.get('author', ''),
        'sections':      data.get('sections', ['summary', 'milestones', 'overdue', 'blocked', 'schedule']),
        'group_by':      data.get('group_by', 'status'),   # none|status|owner|priority
        'filter_status': data.get('filter_status', []),    # [] = all
        'filter_priority': data.get('filter_priority', []),
        'top_level_only': data.get('top_level_only', False),
        'columns':       data.get('columns', ['name', 'owner', 'status', 'priority', 'start_date', 'end_date', 'pct_complete']),
        'max_rows_per_slide': int(data.get('max_rows_per_slide', 12)),
    }

    raw = db.execute("SELECT * FROM tasks WHERE project_id=? ORDER BY sort_order", (pid,)).fetchall()
    all_tasks = [dict(t) for t in raw]

    # Apply filters
    tasks = all_tasks
    if opts['top_level_only']:
        tasks = [t for t in tasks if not t.get('parent_id')]
    if opts['filter_status']:
        tasks = [t for t in tasks if t.get('status') in opts['filter_status']]
    if opts['filter_priority']:
        tasks = [t for t in tasks if t.get('priority') in opts['filter_priority']]

    today = __import__('datetime').date.today().isoformat()
    total  = len(all_tasks)
    done   = sum(1 for t in all_tasks if t.get('status') == 'Complete')
    inprog = sum(1 for t in all_tasks if t.get('status') == 'In Progress')
    blocked_count = sum(1 for t in all_tasks if t.get('status') == 'Blocked')
    overdue_tasks = [t for t in all_tasks if t.get('end_date') and t['end_date'] < today and t.get('status') != 'Complete']
    milestones    = [t for t in all_tasks if t.get('is_milestone') == 1 or t.get('duration') == 0]
    blocked_tasks = [t for t in all_tasks if t.get('status') == 'Blocked']
    health = round(done / total * 100) if total else 0
    health_label = 'On Track' if health >= 70 else ('At Risk' if health >= 40 else 'Off Track')
    health_color = GREEN if health >= 70 else (ORANGE if health >= 40 else RED)

    # Status → color mapping for tags/pips
    status_color_map = {
        'Complete': 'green', 'In Progress': 'blue',
        'Blocked': 'red', 'Not Started': 'gray',
    }
    priority_color_map = {
        'Critical': 'red', 'High': 'orange', 'Medium': 'blue', 'Low': 'gray',
    }

    deck = AppleKeynote(opts['title'])

    # ── Slide 1: Title ────────────────────────────────────────────────────────
    deck.add_title_slide(
        title=opts['title'],
        subtitle_accent=opts['subtitle'],
        description=f"{health_label}  ·  {health}% Complete",
        date=today,
        author=opts['author'],
    )

    # ── Slide 2: Summary Stats ────────────────────────────────────────────────
    if 'summary' in opts['sections']:
        slide, y = deck.add_content_slide(
            'PROJECT HEALTH',
            [(opts['title'], 'white'), (' — Summary', 'secondary')],
        )
        stats = [
            (str(total),         'Total Tasks',    'white'),
            (str(done),          'Complete',       'green'),
            (str(inprog),        'In Progress',    'blue'),
            (str(blocked_count), 'Blocked',        'red'),
            (str(len(overdue_tasks)), 'Overdue',   'orange'),
            (f'{health}%',       health_label,     health_color),
        ]
        sx = MARGIN
        for val, lbl, clr in stats:
            add_stat(slide, sx, y, val, lbl, clr)
            sx += Inches(1.95)
        # Progress bar
        bar_y = y + Inches(1.1)
        bar_w = SLIDE_W - MARGIN * 2
        from apple_keynote import add_rect, add_rounded_rect, NEAR_BLACK, CARD_BORDER
        add_rounded_rect(slide, MARGIN, bar_y, bar_w, Inches(0.18), NEAR_BLACK, CARD_BORDER)
        if health > 0:
            add_rounded_rect(slide, MARGIN, bar_y, bar_w * health / 100, Inches(0.18), health_color)

    # ── Slide 3: Milestones ───────────────────────────────────────────────────
    if 'milestones' in opts['sections'] and milestones:
        slide, y = deck.add_content_slide(
            'MILESTONES',
            [('Key Milestones', 'white')],
        )
        add_section_label(slide, MARGIN, y, f'{len(milestones)} milestones')
        y += Inches(0.3)
        col_w = (SLIDE_W - MARGIN * 2) / 2 - Inches(0.15)
        for i, m in enumerate(milestones[:16]):
            col_x = MARGIN + (i % 2) * (col_w + Inches(0.3))
            row_y = y + (i // 2) * Inches(0.38)
            status_c = status_color_map.get(m.get('status', ''), 'gray')
            date_txt = m.get('end_date') or '—'
            add_text_box(slide, col_x, row_y, col_w - Inches(1.0), Inches(0.25),
                         m.get('name', ''), FONT_BODY, 10, TEXT_PRIMARY)
            add_tag(slide, col_x + col_w - Inches(0.9), row_y + Pt(2), date_txt, status_c)

    # ── Slide 4: Overdue / Blocked ────────────────────────────────────────────
    if 'overdue' in opts['sections'] and (overdue_tasks or blocked_tasks):
        slide, y = deck.add_content_slide(
            'RISKS & BLOCKERS',
            [('Overdue', 'red'), (' & Blocked Tasks', 'secondary')],
        )
        items = [(t, 'red') for t in overdue_tasks[:6]] + [(t, 'orange') for t in blocked_tasks[:4]]
        for t, sev in items[:8]:
            reason = 'Overdue' if sev == 'red' else 'Blocked'
            desc = f"Owner: {t.get('owner') or 'Unassigned'}  ·  Due: {t.get('end_date') or '—'}"
            h = add_risk_card(slide, MARGIN, y,
                              SLIDE_W - MARGIN * 2, t.get('name', ''), desc,
                              severity=sev)
            y += h + Inches(0.12)
            if y > SLIDE_H - Inches(0.8):
                break

    # ── Slides 5+: Schedule table (paginated) ─────────────────────────────────
    if 'schedule' in opts['sections'] and tasks:
        COL_DEFS = {
            'name':        ('Task',     2.8),
            'owner':       ('Owner',    1.3),
            'status':      ('Status',   1.1),
            'priority':    ('Priority', 1.0),
            'start_date':  ('Start',    1.1),
            'end_date':    ('Due',      1.1),
            'pct_complete':('%',        0.55),
            'notes':       ('Notes',    2.5),
        }
        columns = [(COL_DEFS[c][0], COL_DEFS[c][1]) for c in opts['columns'] if c in COL_DEFS]

        def make_cell(t, col_key):
            val = t.get(col_key, '') or ''
            if col_key == 'status':
                return {'text': str(val), 'tag': True, 'tag_color': status_color_map.get(str(val), 'gray')}
            if col_key == 'priority':
                return {'text': str(val), 'tag': True, 'tag_color': priority_color_map.get(str(val), 'gray')}
            if col_key == 'pct_complete':
                return {'text': f'{val}%', 'color': TEXT_SECONDARY}
            return str(val)

        # Group if requested
        if opts['group_by'] != 'none':
            groups = {}
            for t in tasks:
                key = t.get(opts['group_by']) or f'No {opts["group_by"].title()}'
                groups.setdefault(key, []).append(t)
        else:
            groups = {'All Tasks': tasks}

        mxr = opts['max_rows_per_slide']
        for grp_name, grp_tasks in groups.items():
            for page_start in range(0, len(grp_tasks), mxr):
                chunk = grp_tasks[page_start:page_start + mxr]
                label = f'SCHEDULE — {grp_name.upper()}' if opts['group_by'] != 'none' else 'SCHEDULE'
                pg_suffix = f'  ({page_start+1}–{page_start+len(chunk)} of {len(grp_tasks)})' if len(grp_tasks) > mxr else ''
                slide, y = deck.add_content_slide(
                    label,
                    [(grp_name + pg_suffix, 'white')],
                )
                rows = [[make_cell(t, c) for c in opts['columns'] if c in COL_DEFS] for t in chunk]
                add_table_card(slide, MARGIN, y, SLIDE_W - MARGIN * 2, columns, rows)

    # ── Save to buffer ────────────────────────────────────────────────────────
    buf = _io.BytesIO()
    deck.prs.save(buf)
    buf.seek(0)

    safe_name = ''.join(c if c.isalnum() or c in ' -_' else '_' for c in opts['title'])
    from flask import send_file
    return send_file(
        buf, as_attachment=True,
        download_name=f"{safe_name} — Schedule.pptx",
        mimetype='application/vnd.openxmlformats-officedocument.presentationml.presentation',
    )


@app.route('/api/radar/<radar_id>', methods=['PUT'])
def update_radar_fields(radar_id):
    """Update Radar fields via radar.py CLI."""
    data = request.get_json() or {}
    updates = []
    field_map = {
        'title': '--title', 'assignee': '--assignee', 'priority': '--priority',
        'state': '--state', 'description': '--description',
    }
    for field, flag in field_map.items():
        if field in data:
            updates += [flag, str(data[field])]
    if not updates:
        return jsonify({'error': 'No updatable fields provided'}), 400
    try:
        import subprocess
        cmd = ['python3', os.path.join(_ROOT, 'integrations/radar/radar.py'),
               'update_radar', radar_id] + updates
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=15, cwd=_ROOT)
        output = result.stdout.strip()
        start = output.find('{')
        if start >= 0:
            return jsonify(json.loads(output[start:]))
        return jsonify({'ok': True, 'raw': output[:300]})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/projects/<int:pid>/ai-schedule', methods=['POST'])
def ai_schedule(pid):
    """
    Takes a natural-language description (and optional image), calls Floodgate/Claude
    to extract tasks, then creates them in the project.
    Returns {tasks, clarifications}.
    """
    data = request.get_json() or {}
    prompt_text = data.get('prompt', '').strip()
    image_b64 = data.get('image')       # base64 string, no data-URL prefix
    media_type = data.get('media_type', 'image/jpeg')

    if not prompt_text and not image_b64:
        return jsonify({'error': 'No prompt or image provided'}), 400

    # Try Floodgate first, fall back to anthropic SDK
    response_text = _call_llm_for_schedule(prompt_text, image_b64, media_type)
    if isinstance(response_text, dict) and 'error' in response_text:
        return jsonify(response_text), 500

    # Parse JSON from LLM response
    try:
        import re as _re
        # Extract JSON block from response
        json_match = _re.search(r'\{[\s\S]*\}', response_text)
        if not json_match:
            return jsonify({'error': 'LLM did not return parseable JSON', 'raw': response_text}), 500
        parsed = json.loads(json_match.group(0))
    except Exception as e:
        return jsonify({'error': f'JSON parse error: {e}', 'raw': response_text}), 500

    clarifications = parsed.get('clarifications', [])
    raw_tasks = parsed.get('tasks', [])

    if not raw_tasks and not clarifications:
        return jsonify({'error': 'No tasks extracted', 'raw': response_text}), 500

    created = []
    if raw_tasks:
        db = _get_db()
        # First pass: create all tasks to get real IDs
        # We need to map the LLM's local task IDs to real DB IDs
        id_map = {}  # llm_id -> db_id
        max_order = db.execute(
            "SELECT COALESCE(MAX(sort_order),0) FROM tasks WHERE project_id=?", (pid,)
        ).fetchone()[0]

        for i, t in enumerate(raw_tasks):
            llm_id = t.get('id', i + 1)
            cur = db.execute(
                """INSERT INTO tasks
                   (project_id, name, owner, status, priority, start_date, end_date,
                    pct_complete, notes, sort_order, duration)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    pid,
                    t.get('name', 'New Task'),
                    t.get('owner', ''),
                    t.get('status', 'Not Started'),
                    t.get('priority', 'Medium'),
                    t.get('start_date', ''),
                    t.get('end_date', ''),
                    0,
                    t.get('notes', ''),
                    max_order + (i + 1) * 10,
                    t.get('duration', None),
                )
            )
            db.commit()
            real_id = cur.lastrowid
            id_map[llm_id] = real_id

        # Second pass: wire up predecessors using real IDs
        for i, t in enumerate(raw_tasks):
            llm_id = t.get('id', i + 1)
            real_id = id_map[llm_id]
            preds_raw = t.get('depends_on', '')
            if preds_raw:
                # Translate LLM local IDs to real DB IDs
                import re as _re2
                def translate_pred(m):
                    lid = int(m.group(1))
                    typ = m.group(2) or 'FS'
                    lag = m.group(3) or ''
                    rid = id_map.get(lid, lid)
                    return f"{rid}{typ}{lag}"
                translated = _re2.sub(
                    r'(\d+)(FS|SS)?([+-]\d+[dw])?',
                    translate_pred,
                    preds_raw,
                    flags=_re2.IGNORECASE
                )
                db.execute("UPDATE tasks SET depends_on=? WHERE id=?", (translated, real_id))
                db.commit()

            row = db.execute("SELECT * FROM tasks WHERE id=?", (real_id,)).fetchone()
            created.append(dict(row))

    return jsonify({'tasks': created, 'clarifications': clarifications})


def _call_llm_for_schedule(prompt_text, image_b64=None, media_type='image/jpeg'):
    """Call Claude via Floodgate or fallback to anthropic SDK.
    Supports optional vision input: pass image_b64 (base64 string) and media_type.
    """
    system = """You are a project scheduling assistant. Given a description or image of a project,
extract a list of tasks and return ONLY valid JSON in this exact format:

{
  "tasks": [
    {
      "id": 1,
      "name": "Task name",
      "duration": 5,
      "start_date": "",
      "end_date": "",
      "owner": "",
      "priority": "Medium",
      "status": "Not Started",
      "depends_on": "1FS",
      "notes": ""
    }
  ],
  "clarifications": []
}

Rules:
- "id" is a local sequential integer (1, 2, 3...) used only for depends_on references within THIS response
- "duration" is in days (integers). If not specified, make a reasonable estimate.
- "depends_on" uses format "{local_id}FS" or "{local_id}SS" with optional lag like "+3d" or "+1w"
- "start_date" and "end_date": leave empty string "" — the app will compute them from dependencies and duration
- Only set start_date on the very first task(s) with no predecessors, using today's date if not specified
- "priority": one of Low / Medium / High / Critical
- "status": always "Not Started"
- "clarifications": list of strings with questions if critical info is missing (owner, major milestones, etc.)
- If an image is provided, carefully read ALL text and shapes visible in it. Extract every task, phase, step, or milestone you see.
- If anything is ambiguous but you can make a reasonable assumption, DO IT and mention it in notes
- Return ONLY the JSON object, no markdown, no explanation"""

    if image_b64:
        user_content = [
            {'type': 'image', 'source': {'type': 'base64', 'media_type': media_type, 'data': image_b64}},
            {'type': 'text', 'text': f"Build a project schedule from this image.\n\n{prompt_text}" if prompt_text else "Build a complete project schedule from this image. Extract all tasks, phases, milestones, and dependencies you can see."},
        ]
    else:
        user_content = f"Build a project schedule for:\n\n{prompt_text}"

    # Try Floodgate (internal Apple gateway)
    try:
        sys.path.insert(0, _ROOT)
        from integrations.floodgate.floodgate import FloodgateClient
        client = FloodgateClient(model='sonnet')
        result = client.call(
            system=system,
            messages=[{'role': 'user', 'content': user_content}],
            max_tokens=4096
        )
        return result
    except Exception as fg_err:
        fg_msg = str(fg_err)

    # Fallback: Apple proxy at ANTHROPIC_BASE_URL using OIDC token, or real API key
    try:
        import anthropic
        import subprocess as _sp
        base_url = os.environ.get('ANTHROPIC_BASE_URL', 'https://api.anthropic.com')
        api_key = os.environ.get('ANTHROPIC_API_KEY', '')
        # Try to get Apple OIDC token for the local proxy
        if not api_key and 'localhost' in base_url:
            _tok_script = os.path.expanduser('~/.claude/apple/get-apple-token.sh')
            if os.path.exists(_tok_script):
                try:
                    _tok = _sp.run(['bash', _tok_script], capture_output=True, text=True, timeout=10)
                    api_key = _tok.stdout.strip() or 'no-key'
                except Exception:
                    api_key = 'no-key'
        if not api_key:
            api_key = 'no-key'
        client = anthropic.Anthropic(base_url=base_url, api_key=api_key)
        msg = client.messages.create(
            model='claude-sonnet-4-6',
            max_tokens=4096,
            system=system,
            messages=[{'role': 'user', 'content': user_content}]
        )
        return msg.content[0].text
    except Exception as e:
        return {'error': f'LLM unavailable. Floodgate: {fg_msg[:80]}. Anthropic: {e}.'}


# ── DB init helper ────────────────────────────────────────────────────────────
def _init_db():
    db = sqlite3.connect(app.config['DATABASE'])
    db.execute("PRAGMA foreign_keys = ON")
    db.executescript("""
        CREATE TABLE IF NOT EXISTS projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            color TEXT DEFAULT '#4A90E2',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
            name TEXT NOT NULL DEFAULT 'New Task',
            owner TEXT DEFAULT '',
            status TEXT DEFAULT 'Not Started',
            priority TEXT DEFAULT 'Medium',
            start_date TEXT DEFAULT '',
            end_date TEXT DEFAULT '',
            pct_complete INTEGER DEFAULT 0,
            depends_on TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            sort_order INTEGER DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS comments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
            text TEXT NOT NULL,
            author TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
    """)
    db.commit()
    count = db.execute("SELECT COUNT(*) FROM projects").fetchone()[0]
    if count == 0:
        db.execute(
            "INSERT INTO projects (name, color) VALUES (?, ?)",
            ("My First Project", "#4A90E2")
        )
        db.commit()
    # Migrations: add columns added after initial schema
    for col, defn in [
        ('duration',        'INTEGER DEFAULT NULL'),
        ('radar_id',        'TEXT DEFAULT ""'),
        ('is_milestone',    'INTEGER DEFAULT 0'),
        ('parent_id',       'INTEGER DEFAULT NULL'),
        ('baseline_start',  'TEXT DEFAULT ""'),
        ('baseline_end',    'TEXT DEFAULT ""'),
        ('color',           'TEXT DEFAULT ""'),
    ]:
        try:
            db.execute(f"ALTER TABLE tasks ADD COLUMN {col} {defn}")
            db.commit()
        except sqlite3.OperationalError:
            pass  # column already exists
    # Users + project_access tables
    db.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            apple_sub TEXT UNIQUE DEFAULT '',
            email TEXT DEFAULT '',
            full_name TEXT DEFAULT '',
            username TEXT DEFAULT '',
            password_hash TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS project_access (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
            user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            role TEXT DEFAULT 'editor',
            UNIQUE(project_id, user_id)
        );
    """)
    db.commit()
    # Migrate users table — add apple_sub / full_name if they don't exist yet
    for col, defn in [
        ('apple_sub', 'TEXT DEFAULT ""'),
        ('full_name', 'TEXT DEFAULT ""'),
    ]:
        try:
            db.execute(f"ALTER TABLE users ADD COLUMN {col} {defn}")
            db.commit()
        except sqlite3.OperationalError:
            pass
    # Migrate projects table — add user_id
    for col, defn in [('user_id', 'INTEGER DEFAULT NULL')]:
        try:
            db.execute(f"ALTER TABLE projects ADD COLUMN {col} {defn}")
            db.commit()
        except sqlite3.OperationalError:
            pass
    db.close()


def _hash_password(pw):
    salt = secrets.token_hex(16)
    h = hashlib.sha256((salt + pw).encode()).hexdigest()
    return f"{salt}:{h}"

def _check_password(pw, stored):
    try:
        salt, h = stored.split(':', 1)
        return hashlib.sha256((salt + pw).encode()).hexdigest() == h
    except Exception:
        return False


# ── Project Members ─────────────────────────────────────────────────────────────
@app.route('/api/projects/<int:pid>/members', methods=['GET'])
def list_project_members(pid):
    db = _get_db()
    rows = db.execute("""
        SELECT pa.user_id, u.username, pa.role
        FROM project_access pa JOIN users u ON u.id = pa.user_id
        WHERE pa.project_id=?
        ORDER BY pa.role, u.username
    """, (pid,)).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/projects/<int:pid>/members', methods=['POST'])
def add_project_member(pid):
    data = request.get_json() or {}
    username = (data.get('username') or '').strip()
    role = data.get('role', 'editor')
    if role not in ('editor', 'viewer', 'owner'):
        role = 'editor'
    db = _get_db()
    user = db.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone()
    if not user:
        return jsonify({'error': f'User "{username}" not found'}), 404
    try:
        db.execute("INSERT OR REPLACE INTO project_access (project_id, user_id, role) VALUES (?,?,?)",
                   (pid, user['id'], role))
        db.commit()
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    return jsonify({'ok': True})

@app.route('/api/projects/<int:pid>/members/<int:uid>', methods=['DELETE'])
def remove_project_member(pid, uid):
    db = _get_db()
    db.execute("DELETE FROM project_access WHERE project_id=? AND user_id=?", (pid, uid))
    db.commit()
    return jsonify({'ok': True})


# ── CSV Import ──────────────────────────────────────────────────────────────────
@app.route('/api/projects/<int:pid>/import/csv', methods=['POST'])
def import_csv(pid):
    """Accept parsed CSV rows as JSON (from client-side parsing)."""
    data = request.get_json()
    if not data or not isinstance(data, list):
        return jsonify({'error': 'Expected JSON array of rows'}), 400

    db = _get_db()
    proj = db.execute("SELECT id FROM projects WHERE id=?", (pid,)).fetchone()
    if not proj:
        return jsonify({'error': 'Project not found'}), 404

    VALID_STATUS   = {'Not Started','In Progress','Complete','Blocked'}
    VALID_PRIORITY = {'Low','Medium','High','Critical'}

    created = updated = skipped = 0
    row_ref_to_id = {}  # row_ref → DB id
    deferred_parents = []  # (db_id, parent_row_ref)
    deferred_preds   = []  # (db_id, pred_str)

    for r in data:
        name = str(r.get('name') or '').strip()
        if not name:
            skipped += 1
            continue
        status   = r.get('status',   'Not Started')
        priority = r.get('priority', 'Medium')
        if status   not in VALID_STATUS:   status   = 'Not Started'
        if priority not in VALID_PRIORITY: priority = 'Medium'
        try: duration = int(r.get('duration') or 0) or None
        except: duration = None
        try: pct = max(0, min(100, int(r.get('pct') or 0)))
        except: pct = 0
        row_ref  = r.get('row_ref')
        task_id  = r.get('task_id')
        pred_str = str(r.get('predecessor') or '').strip()
        parent_ref = str(r.get('parent_row') or '').strip()

        if task_id:
            existing = db.execute("SELECT id FROM tasks WHERE id=? AND project_id=?", (task_id, pid)).fetchone()
        else:
            existing = None

        payload = dict(name=name, owner=str(r.get('owner') or ''), status=status, priority=priority,
                       start_date=str(r.get('start_date') or ''), end_date=str(r.get('end_date') or ''),
                       pct_complete=pct, notes=str(r.get('notes') or ''),
                       radar_id=str(r.get('radar_id') or ''))
        if duration is not None:
            payload['duration'] = duration

        if existing:
            set_clause = ', '.join(f"{k}=?" for k in payload)
            db.execute(f"UPDATE tasks SET {set_clause} WHERE id=?", list(payload.values()) + [task_id])
            db_id = task_id
            updated += 1
        else:
            max_order = db.execute("SELECT COALESCE(MAX(sort_order),0) FROM tasks WHERE project_id=?", (pid,)).fetchone()[0]
            cur = db.execute(
                """INSERT INTO tasks (project_id, name, owner, status, priority,
                   start_date, end_date, pct_complete, notes, sort_order)
                   VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (pid, name, payload['owner'], status, priority,
                 payload['start_date'], payload['end_date'], pct, payload['notes'], max_order + 10)
            )
            db_id = cur.lastrowid
            created += 1

        if row_ref:
            row_ref_to_id[str(row_ref)] = db_id
        if parent_ref:
            deferred_parents.append((db_id, parent_ref))
        if pred_str:
            deferred_preds.append((db_id, pred_str))

    db.commit()

    # Wire parent_id
    for db_id, parent_ref in deferred_parents:
        parent_db_id = row_ref_to_id.get(str(parent_ref))
        if parent_db_id:
            db.execute("UPDATE tasks SET parent_id=? WHERE id=?", (parent_db_id, db_id))

    # Wire depends_on (predecessor refs are row numbers separated by ;)
    import re as _re
    for db_id, pred_str in deferred_preds:
        stored_parts = []
        for part in _re.split(r'[;,]', pred_str):
            m = _re.match(r'^(\d+)(FS|SS|FF|SF)?([+-]\d+[dw])?$', part.strip(), _re.I)
            if m:
                ref_db_id = row_ref_to_id.get(m.group(1))
                if ref_db_id:
                    stored_parts.append(f"{ref_db_id}{(m.group(2) or 'FS').upper()}{m.group(3) or ''}")
        if stored_parts:
            db.execute("UPDATE tasks SET depends_on=? WHERE id=?", (','.join(stored_parts), db_id))

    db.commit()
    return jsonify({'ok': True, 'created': created, 'updated': updated, 'skipped': skipped})


# ── Comments ───────────────────────────────────────────────────────────────────
@app.route('/api/tasks/<int:tid>/comments', methods=['GET'])
def get_comments(tid):
    db = _get_db()
    rows = db.execute(
        "SELECT * FROM comments WHERE task_id=? ORDER BY created_at", (tid,)
    ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/tasks/<int:tid>/comments', methods=['POST'])
def add_comment(tid):
    data = request.json or {}
    if not data.get('text','').strip():
        return jsonify({'error': 'text required'}), 400
    db = _get_db()
    cur = db.execute(
        "INSERT INTO comments (task_id, text, author) VALUES (?,?,?)",
        (tid, data['text'].strip(), data.get('author',''))
    )
    db.commit()
    row = db.execute("SELECT * FROM comments WHERE id=?", (cur.lastrowid,)).fetchone()
    return jsonify(dict(row))

@app.route('/api/comments/<int:cid>', methods=['DELETE'])
def delete_comment(cid):
    db = _get_db()
    db.execute("DELETE FROM comments WHERE id=?", (cid,))
    db.commit()
    return jsonify({'ok': True})


# ── Baseline ───────────────────────────────────────────────────────────────────
@app.route('/api/projects/<int:pid>/baseline', methods=['POST'])
def set_baseline(pid):
    db = _get_db()
    tasks = db.execute(
        "SELECT id, start_date, end_date FROM tasks WHERE project_id=?", (pid,)
    ).fetchall()
    for t in tasks:
        db.execute(
            "UPDATE tasks SET baseline_start=?, baseline_end=? WHERE id=?",
            (t['start_date'] or '', t['end_date'] or '', t['id'])
        )
    db.commit()
    return jsonify({'ok': True, 'count': len(tasks)})

@app.route('/api/projects/<int:pid>/baseline', methods=['DELETE'])
def clear_baseline(pid):
    db = _get_db()
    db.execute("UPDATE tasks SET baseline_start='', baseline_end='' WHERE project_id=?", (pid,))
    db.commit()
    return jsonify({'ok': True})


if __name__ == '__main__':
    _init_db()
    print("Project Manager running at http://localhost:5000")
    app.run(debug=True, host='0.0.0.0', port=5000, use_reloader=False)
